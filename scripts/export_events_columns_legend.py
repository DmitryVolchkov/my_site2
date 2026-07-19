#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Создаёт Excel-справочник по колонкам data/events.csv"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent.parent / 'data' / 'events_columns_legend.xlsx'

ROWS = [
    ('id', 'да*', 'Уникальный код события (например ev-0001). Переход таймлайна по дате, карточка факта. В JSON → unique_id.'),
    ('start_year', 'да', 'Год начала. Без года строка не попадёт в таймлайн.'),
    ('start_month', 'нет', 'Месяц начала (1–12). Пусто — только год.'),
    ('start_day', 'нет', 'День начала (1–31). Обычно вместе с start_month.'),
    ('end_year', 'нет', 'Год окончания — для длительных событий.'),
    ('end_month', 'нет', 'Месяц окончания периода.'),
    ('end_day', 'нет', 'День окончания. При end_year на шкале может быть полоса периода.'),
    ('headline', 'желательно', 'Заголовок на маркере и в блоке факта.'),
    ('text', 'желательно', 'Описание события.'),
    ('media_url', 'нет', 'Ссылка на изображение, видео или страницу.'),
    ('media_caption', 'нет', 'Подпись к медиа.'),
    ('media_credit', 'нет', 'Источник или автор медиа.'),
    ('group', 'нет', 'Полоса на таймлайне (например «История»).'),
    ('tags', 'нет', 'Теги через «;» или «,». В JSON → _tags.'),
    ('importance', 'нет', 'Число 1–10: важность для scale_data.json.'),
]

EXAMPLE_HEADERS = [
    'id', 'start_year', 'start_month', 'start_day', 'end_year', 'end_month', 'end_day',
    'headline', 'text', 'media_url', 'media_caption', 'media_credit', 'group', 'tags', 'importance',
]
EXAMPLE_ROW = [
    'ev-0001', 1945, 5, 9, None, None, None,
    'День Победы', 'Окончание Великой Отечественной войны в Европе.',
    None, None, None, 'История', 'СССР;Праздники', 3,
]

MAPPING = [
    ('id', 'unique_id'),
    ('start_year, start_month, start_day', 'start_date { year, month?, day? }'),
    ('end_year, end_month, end_day', 'end_date (если задан год окончания)'),
    ('headline', 'text.headline'),
    ('text', 'text.text'),
    ('media_url, media_caption, media_credit', 'media { url, caption?, credit? }'),
    ('group', 'group'),
    ('tags', '_tags (массив)'),
    ('importance', '_importance'),
]


def style_header(cell):
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor='1A2A3A')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = _border()


def _border():
    thin = Side(style='thin', color='D4CFC4')
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def main():
    wb = Workbook()
    wrap = Alignment(wrap_text=True, vertical='top')

    ws = wb.active
    ws.title = 'Расшифровка колонок'
    for col, h in enumerate(['Поле', 'Обязательно', 'Что означает'], 1):
        style_header(ws.cell(row=1, column=col, value=h))
    for r, row in enumerate(ROWS, 2):
        for col, val in enumerate(row, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.alignment = wrap
            c.border = _border()
    note = len(ROWS) + 3
    ws.cell(row=note, column=1, value='Примечание').font = Font(bold=True)
    ws.merge_cells(start_row=note, start_column=2, end_row=note, end_column=3)
    ws.cell(
        row=note, column=2,
        value='* id должен быть уникальным. ev-0004 сейчас дублируется — лучше ev-0004 и ev-0005.',
    ).alignment = wrap
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 72
    ws.freeze_panes = 'A2'

    ws2 = wb.create_sheet('Пример строки')
    for col, h in enumerate(EXAMPLE_HEADERS, 1):
        style_header(ws2.cell(row=1, column=col, value=h))
    for col, v in enumerate(EXAMPLE_ROW, 1):
        c = ws2.cell(row=2, column=col, value=v)
        c.alignment = wrap
        c.border = _border()
    for i in range(1, len(EXAMPLE_HEADERS) + 1):
        ws2.column_dimensions[get_column_letter(i)].width = 14
    ws2.freeze_panes = 'A2'

    ws3 = wb.create_sheet('CSV → JSON')
    for col, h in enumerate(['Поле в CSV', 'Поле в timeline_data.json'], 1):
        style_header(ws3.cell(row=1, column=col, value=h))
    for r, row in enumerate(MAPPING, 2):
        for col, val in enumerate(row, 1):
            c = ws3.cell(row=r, column=col, value=val)
            c.alignment = wrap
            c.border = _border()
    ws3.column_dimensions['A'].width = 36
    ws3.column_dimensions['B'].width = 40
    ws3.freeze_panes = 'A2'

    ws4 = wb.create_sheet('Конвертация')
    for r, (a, b) in enumerate([
        ('Команда', 'python scripts/convert_csv_to_timelinejson.py data/events.csv --out-dir ./data'),
        ('Результат', 'data/timeline_data.json — для TimelineJS'),
        ('', 'data/scale_data.json — вспомогательные данные шкалы'),
        ('Источник', 'data/events.csv'),
    ], 1):
        if a:
            ws4.cell(row=r, column=1, value=a).font = Font(bold=True)
        ws4.cell(row=r, column=2, value=b).alignment = wrap
    ws4.column_dimensions['A'].width = 14
    ws4.column_dimensions['B'].width = 70

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print('OK:', OUT)


if __name__ == '__main__':
    main()
